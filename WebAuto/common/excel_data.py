from xlrd import open_workbook
import xlwt
import openpyxl
import os



class SheetTypeError(Exception):
    pass

class SheetNotFoundError(Exception):
    pass

class SheetIsNullError(Exception):
    pass

class Excel_Data:

    def __init__(self,excel_path,sheet="Sheet1",clear=True):
        if os.path.exists(excel_path):
            self.excel=excel_path
        else:
            raise FileNotFoundError("{0}文件不存在！".format(excel_path))
        self._data=list()
        self.sheet=sheet
        self.clear_data=clear


    @property
    #方法一：xlrd读取excel，速度更快
    #数据为列表类型数据，每个值是一个字典
    def data(self):
        if not self._data:
            workbook=open_workbook(self.excel)

            if type(self.sheet) not in [int,str]:
                raise SheetTypeError("please pass in <type int> or <type str>, not {0} ",format(type(self.sheet)))
            elif type(self.sheet)==int:
                try:
                    ws=workbook.sheet_by_index(self.sheet)
                except:
                    raise SheetNotFoundError("{0} 不存在", format(self.sheet))
            else:
                try:
                    ws=workbook.sheet_by_name(self.sheet)
                except:
                    raise SheetNotFoundError("{0} 不存在", format(self.sheet))
            if ws.nrows>=1:
                title=ws.row_values(0)
                for col in range(1, ws.nrows):
                    self._data.append(dict(zip(title, ws.row_values(col))))
            else:
                raise SheetIsNullError("sheet 为空")
        return self._data

    # 方法二：openpyxl读取excel
    # def data(self):
    #     if not self._data:
    #         excel=openpyxl.load_workbook(self.excel)
    #         if type(self.sheet) not in [int,str]:
    #             raise SheetTypeError("please pass in type<int> or type<str> , not {0}",format(type(self.sheet)))
    #         elif type(self.sheet)==int:
    #             try:
    #                 ws=excel.get_sheet_by_name(excel.get_sheet_names[0])
    #             except:
    #                 raise SheetNotFoundError("{0} 不存在",format(self.sheet))
    #         else:
    #             try:
    #                 ws=excel[self.sheet]
    #             except:
    #                 raise SheetNotFoundError("{0} 不存在",format(self.sheet))
    #         if ws.max_row>=1:
    #             title=[cell.value for cell in list(ws.rows)[0]]
    #             for col in range(1,s.max_row):
    #                 value_col=[c.value for c in list(ws.rows)[col]]
    #                 self._data.append(dict(zip(title,value_col)))
    #         else:
    #             raise SheetIsNullError("{} 为空",format(self.sheet))
    #     return self._data



    @data.setter
    def data(self,param):
        """单行写入字典类型的数据"""
        excel=openpyxl.load_workbook(self.excel)
        if type(self.sheet) != str:
            raise SheetTypeError("please pass in type<string>,not {0}",format(type(self.sheet)))
        else:
            try:
                ws =excel[self.sheet]
            except:
                ws =excel.create_sheet(self.sheet)
            try:
                if self.clear_data==True :
                    del excel[self.sheet]
                    ws = excel.create_sheet(self.sheet)
                    self.clear_data=False
            except:
                pass
        i = 1
        row=ws.max_row
        for k,v in (param.get("value") or param).items():
            ws.cell(row=1,column=i).value=k
            ws.cell(row=row+1,column=i).value=str(v)
            i=i+1
        excel.save(self.excel)












